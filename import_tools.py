import csv
from io import BytesIO, StringIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

import storage

COLUMNS = [
    ("plate", "Plate"),
    ("driver_name", "Driver Name"),
    ("phone", "Phone"),
    ("unit", "Unit / Department"),
    ("company", "Company"),
    ("employee_code", "Employee Code"),
    ("vehicle_model", "Vehicle Model"),
    ("vehicle_color", "Vehicle Color"),
    ("status", "Status"),
]

SAMPLE = [
    "12A34567",
    "John Doe",
    "+0000000000",
    "Security",
    "Example Company",
    "EMP-001",
    "Sedan",
    "White",
    "allowed",
]


def _normalize_header(value):
    return str(value or "").strip().lower().replace(" ", "_").replace("/", "_").replace("-", "_")


def _field_for_header(value):
    normalized = _normalize_header(value)
    for key, label in COLUMNS:
        aliases = {key, _normalize_header(label), _normalize_header(label.replace(" / ", " "))}
        if normalized in aliases:
            return key
    extra_aliases = {
        "department": "unit",
        "unit_department": "unit",
        "driver_phone": "phone",
        "car_model": "vehicle_model",
        "car_color": "vehicle_color",
        "plate_number": "plate",
    }
    return extra_aliases.get(normalized)


def build_vehicle_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "vehicles"
    headers = [label for _, label in COLUMNS]
    ws.append(headers)
    ws.append(SAMPLE)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    widths = [18, 24, 18, 24, 20, 18, 20, 18, 16]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    ws.freeze_panes = "A2"
    notes = wb.create_sheet("README")
    notes.append(["Column", "Description"])
    notes.append(["Plate", "Required. Unique per site. Example: 12A34567"])
    notes.append(["Status", "allowed, review, unknown, blocked"])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def build_vehicle_csv_template():
    stream = StringIO()
    writer = csv.writer(stream)
    writer.writerow([label for _, label in COLUMNS])
    writer.writerow(SAMPLE)
    output = BytesIO()
    output.write("\ufeff".encode("utf-8"))
    output.write(stream.getvalue().encode("utf-8"))
    output.seek(0)
    return output


def read_vehicle_excel(file_storage):
    wb = load_workbook(file_storage, data_only=True)
    ws = wb.active
    header = [str(c.value or "").strip() for c in ws[1]]
    mapping = {}
    for idx, value in enumerate(header):
        key = _field_for_header(value)
        if key:
            mapping[idx] = key
    rows = []
    for number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        item = {"_row": number}
        for idx, value in enumerate(row):
            key = mapping.get(idx)
            if key:
                item[key] = str(value or "").strip()
        if item.get("plate"):
            rows.append(item)
    return rows


def read_vehicle_csv(file_storage):
    raw = file_storage.read()
    if isinstance(raw, bytes):
        text = raw.decode("utf-8-sig")
    else:
        text = str(raw or "")
    reader = csv.DictReader(StringIO(text))
    rows = []
    for number, row in enumerate(reader, start=2):
        item = {"_row": number}
        for header, value in row.items():
            key = _field_for_header(header)
            if key:
                item[key] = str(value or "").strip()
        if item.get("plate"):
            rows.append(item)
    return rows


def _import_rows(rows, default_site_id=None):
    result = {"imported": 0, "errors": [], "valid": 0, "duplicates_in_file": 0}
    seen = set()
    for item in rows:
        row_num = item.pop("_row", "-")
        plate = storage.clean_plate(item.get("plate"))
        if not plate:
            result["errors"].append({"row": row_num, "error": "Plate is required"})
            continue
        if plate in seen:
            result["duplicates_in_file"] += 1
        seen.add(plate)
        try:
            item["site_id"] = default_site_id
            item["plate"] = plate
            storage.save_vehicle(item)
            result["imported"] += 1
            result["valid"] += 1
        except Exception as exc:
            result["errors"].append({"row": row_num, "error": str(exc)})
    return result


def import_vehicles_excel(file_storage, default_site_id=None):
    if not file_storage:
        return {"imported": 0, "errors": [{"row": "-", "error": "No file uploaded"}]}
    try:
        rows = read_vehicle_excel(file_storage)
    except Exception as exc:
        return {"imported": 0, "errors": [{"row": "-", "error": f"Could not read Excel file: {exc}"}]}
    return _import_rows(rows, default_site_id)


def import_vehicles_csv(file_storage, default_site_id=None):
    if not file_storage:
        return {"imported": 0, "errors": [{"row": "-", "error": "No file uploaded"}]}
    try:
        rows = read_vehicle_csv(file_storage)
    except Exception as exc:
        return {"imported": 0, "errors": [{"row": "-", "error": f"Could not read CSV file: {exc}"}]}
    return _import_rows(rows, default_site_id)
